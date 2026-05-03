from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from gestion_camiones.data.models import ViajeResumen
from gestion_camiones.services.report_exporter import (
    build_monthly_report_summary,
    build_ricco_export_filename,
    export_monthly_report_excel,
    export_monthly_report_pdf,
    export_ricco_report_excel,
)


def _sample_viaje() -> ViajeResumen:
    return ViajeResumen(
        id=1,
        fecha="2026-04-15",
        cliente="Cliente Uno",
        carta_porte="CP-1001",
        carga="CONT-001",
        lugar_carga="Buenos Aires",
        lugar_descarga="Cordoba",
        observaciones="",
        chofer="Juan Perez",
        tipo_carga="General",
        camion="Camion 1 - AAA111",
        semi="",
        tarifa=100000,
        fecha_descarga_tarifa="2026-04-16",
        hay_demora=True,
        demora=5000,
        fecha_descarga_demora="2026-04-16",
        descarga_vacio=True,
        vacio=2500,
        fecha_descarga_vacio="2026-04-16",
        lugar_descarga_vacio="Taller",
        gas_oil_lts=320.5,
        peajes=1500,
        estado="Programado",
    )


class ReportExporterTests(unittest.TestCase):
    def test_build_summary(self) -> None:
        summary = build_monthly_report_summary([_sample_viaje(), _sample_viaje()])
        self.assertEqual(summary.trip_count, 2)
        self.assertEqual(summary.total_amount, 218000)

    def test_export_excel(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "reporte.xlsx"
            export_monthly_report_excel(output_path, "Abril 2026", [_sample_viaje()])

            workbook = load_workbook(output_path)
            worksheet = workbook.active

            self.assertEqual(worksheet["A1"].value, "Reporte mensual")
            self.assertEqual(worksheet["A2"].value, "Periodo: Abril 2026")
            self.assertEqual(worksheet["A6"].value, "2026-04-15")
            self.assertEqual(worksheet["C6"].value, "CP-1001")
            self.assertEqual(worksheet["N6"].value, "$ 109.000")

    def test_export_pdf(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "reporte.pdf"
            export_monthly_report_pdf(output_path, "Abril 2026", [_sample_viaje()])

            content = output_path.read_bytes()
            self.assertTrue(content.startswith(b"%PDF"))

    def test_export_ricco_template_excel(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            template_path = Path(temp_dir) / "RICCO.xlsx"
            output_path = Path(temp_dir) / "salida.xlsx"
            self._build_ricco_template(template_path)

            viaje = _sample_viaje()
            viaje = ViajeResumen(
                **{
                    **viaje.__dict__,
                    "cliente": "Cliente Uno (Ricco)",
                    "observaciones": "Ok",
                    "gas_oil_lts": 120.0,
                }
            )
            export_ricco_report_excel(
                template_path,
                output_path,
                "2026-04-01",
                "2026-04-30",
                [viaje],
                company_name="Mi Empresa SRL",
            )

            workbook = load_workbook(output_path)
            worksheet = workbook.active

            self.assertEqual(worksheet["D8"].value, "Mi Empresa SRL")
            self.assertEqual(worksheet["E12"].value.strftime("%Y-%m-%d"), "2026-04-01")
            self.assertEqual(worksheet["H12"].value.strftime("%Y-%m-%d"), "2026-04-30")
            self.assertEqual(worksheet["C17"].value, "CP-1001")
            self.assertEqual(worksheet["E17"].value, "Cliente Uno")
            self.assertEqual(worksheet["F17"].value, "gral")
            self.assertEqual(worksheet["K17"].value, 100000)
            self.assertEqual(worksheet["M17"].value, 7500)
            self.assertEqual(worksheet["N17"].value, 109000)
            self.assertEqual(worksheet["Q17"].value, 109000)
            self.assertEqual(worksheet["R17"].value, "Ok")

    def test_export_ricco_uses_imo_for_dangerous_cargo(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            template_path = Path(temp_dir) / "RICCO.xlsx"
            output_path = Path(temp_dir) / "salida.xlsx"
            self._build_ricco_template(template_path)

            viaje = ViajeResumen(
                **{
                    **_sample_viaje().__dict__,
                    "tipo_carga": "Carga peligrosa",
                }
            )
            export_ricco_report_excel(
                template_path,
                output_path,
                "2026-04-01",
                "2026-04-30",
                [viaje],
                company_name="Mi Empresa SRL",
            )

            workbook = load_workbook(output_path)
            worksheet = workbook.active

            self.assertEqual(worksheet["F17"].value, "imo")

    def test_export_ricco_uses_imo_for_imo_cargo(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            template_path = Path(temp_dir) / "RICCO.xlsx"
            output_path = Path(temp_dir) / "salida.xlsx"
            self._build_ricco_template(template_path)

            viaje = ViajeResumen(
                **{
                    **_sample_viaje().__dict__,
                    "tipo_carga": "IMO",
                }
            )
            export_ricco_report_excel(
                template_path,
                output_path,
                "2026-04-01",
                "2026-04-30",
                [viaje],
                company_name="Mi Empresa SRL",
            )

            workbook = load_workbook(output_path)
            worksheet = workbook.active

            self.assertEqual(worksheet["F17"].value, "imo")

    def test_export_ricco_demora_column_sums_delay_and_empty_trip(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            template_path = Path(temp_dir) / "RICCO.xlsx"
            output_path = Path(temp_dir) / "salida.xlsx"
            self._build_ricco_template(template_path)

            base = _sample_viaje()
            viajes = [
                ViajeResumen(
                    **{
                        **base.__dict__,
                        "id": 1,
                        "demora": 4000,
                        "vacio": 0,
                    }
                ),
                ViajeResumen(
                    **{
                        **base.__dict__,
                        "id": 2,
                        "demora": 0,
                        "vacio": 3000,
                    }
                ),
                ViajeResumen(
                    **{
                        **base.__dict__,
                        "id": 3,
                        "demora": 4000,
                        "vacio": 3000,
                    }
                ),
            ]
            export_ricco_report_excel(
                template_path,
                output_path,
                "2026-04-01",
                "2026-04-30",
                viajes,
                company_name="Mi Empresa SRL",
            )

            workbook = load_workbook(output_path)
            worksheet = workbook.active

            self.assertEqual(worksheet["M17"].value, 4000)
            self.assertEqual(worksheet["M18"].value, 3000)
            self.assertEqual(worksheet["M19"].value, 7000)

    def test_build_ricco_export_filename(self) -> None:
        filename = build_ricco_export_filename(
            "2026-04-01",
            "2026-04-30",
            'Mi Empresa: SRL / Casa Matriz',
        )
        self.assertEqual(
            filename,
            "ricco 01 de abril de 2026 al 30 de abril de 2026 Mi Empresa SRL Casa Matriz.xlsx",
        )

    def _build_ricco_template(self, template_path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Hoja1"

        worksheet["B2"] = "RESUMEN DE PRESTACION DE SERVICIOS"
        worksheet["B6"] = "Cliente:"
        worksheet["D6"] = "Transportes Ricco"
        worksheet["B8"] = "Transportista:"
        worksheet["D8"] = "Plantilla Base"
        worksheet["B10"] = "Resumen N°"
        worksheet["D10"] = "BASE"
        worksheet["B12"] = "Periodo a facturar:"
        worksheet["D12"] = "Desde:"
        worksheet["G12"] = "Hasta:"

        headers = [
            "Fecha de viaje",
            "N° Carta de Porte",
            "Descrip. de flete",
            "Cliente de Ricco",
            "IMO-GRAL",
            "Origen",
            "Destino",
            "Chofer",
            "Patente Equipo",
            "Precio Flete",
            "F.DESC VACIO",
            "Demora",
            "Monto total",
            "Anticipo ( en $ )",
            "Gas Oil ( lrs )",
            "Saldo a cobrar",
            "Comentario",
        ]
        for column_index, value in enumerate(headers, start=2):
            worksheet.cell(16, column_index).value = value
        for row_index in range(17, 37):
            for column_index in range(2, 19):
                worksheet.cell(row_index, column_index).number_format = "General"
        worksheet["B36"] = "Total a facturar"

        workbook.save(template_path)


if __name__ == "__main__":
    unittest.main()
