from __future__ import annotations

import sys
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from gestion_camiones.data.models import ViajeResumen


@dataclass(frozen=True)
class MonthlyReportSummary:
    trip_count: int
    total_amount: float


MONTH_NAMES_LOWER = (
    "enero",
    "febrero",
    "marzo",
    "abril",
    "mayo",
    "junio",
    "julio",
    "agosto",
    "septiembre",
    "octubre",
    "noviembre",
    "diciembre",
)

RICCO_TEMPLATE_NAME = "RICCO.xlsx"
RICCO_DATA_START_ROW = 17
RICCO_TEMPLATE_TOTAL_ROW = 36


def build_monthly_report_summary(viajes: list[ViajeResumen]) -> MonthlyReportSummary:
    return MonthlyReportSummary(
        trip_count=len(viajes),
        total_amount=sum(viaje.costo_total for viaje in viajes),
    )


def export_monthly_report_excel(
    output_path: Path,
    period_label: str,
    viajes: list[ViajeResumen],
    *,
    report_title: str = "Reporte mensual",
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte mensual"

    summary = build_monthly_report_summary(viajes)
    headers = _report_headers()

    worksheet["A1"] = report_title
    worksheet["A1"].font = Font(size=16, bold=True)
    worksheet["A2"] = f"Periodo: {period_label}"
    worksheet["A3"] = f"Viajes: {summary.trip_count}"
    worksheet["D3"] = f"Total: {_format_money(summary.total_amount)}"

    header_fill = PatternFill(fill_type="solid", fgColor="24333A")
    header_font = Font(color="FFFFFF", bold=True)

    header_row = 5
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(_report_rows(viajes), start=header_row + 1):
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(vertical="top")

    widths = [12, 22, 18, 16, 18, 18, 18, 18, 16, 16, 16, 16, 16, 16, 14]
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + column_index)].width = width

    workbook.save(output_path)


def export_monthly_report_pdf(
    output_path: Path,
    period_label: str,
    viajes: list[ViajeResumen],
    *,
    report_title: str = "Reporte mensual",
) -> None:
    summary = build_monthly_report_summary(viajes)
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    story = [
        Paragraph(report_title, styles["Title"]),
        Paragraph(f"Periodo: {period_label}", styles["Heading3"]),
        Paragraph(
            f"Viajes: {summary.trip_count} | Total: {_format_money(summary.total_amount)}",
            styles["BodyText"],
        ),
        Spacer(1, 8),
    ]

    table_data = [_report_headers(), *_report_rows(viajes)]
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#24333A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9E0E5")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFB")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (8, 1), (13, -1), "RIGHT"),
            ]
        )
    )
    story.append(table)
    document.build(story)


def find_ricco_template_path() -> Path | None:
    internal_template = _internal_ricco_template_path()
    if internal_template is not None:
        return internal_template

    candidates = (
        Path.home() / "Desktop" / RICCO_TEMPLATE_NAME,
        Path.home() / "Escritorio" / RICCO_TEMPLATE_NAME,
        Path.home() / "OneDrive" / "Desktop" / RICCO_TEMPLATE_NAME,
        Path.home() / "OneDrive" / "Escritorio" / RICCO_TEMPLATE_NAME,
    )
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def build_ricco_export_filename(
    period_start: str,
    period_end: str,
    company_name: str,
) -> str:
    period_label = (
        f"{_format_long_date(period_start)} al {_format_long_date(period_end)}"
    )
    safe_company = _safe_filename_component(company_name)
    safe_period = _safe_filename_component(period_label)
    return f"ricco {safe_period} {safe_company}.xlsx".strip()


def export_ricco_report_excel(
    template_path: Path,
    output_path: Path,
    period_start: str,
    period_end: str,
    viajes: list[ViajeResumen],
    *,
    company_name: str,
) -> None:
    workbook = load_workbook(template_path)
    worksheet = workbook[workbook.sheetnames[0]]

    extra_rows = max(
        0,
        len(viajes) - (RICCO_TEMPLATE_TOTAL_ROW - RICCO_DATA_START_ROW),
    )
    total_row = RICCO_TEMPLATE_TOTAL_ROW
    if extra_rows:
        worksheet.insert_rows(total_row, extra_rows)
        total_row += extra_rows

    for row_index in range(RICCO_DATA_START_ROW, total_row):
        _copy_row_style(worksheet, RICCO_DATA_START_ROW, row_index)
        _clear_row_values(worksheet, row_index)

    worksheet["D8"] = company_name
    worksheet["E12"] = _excel_date(period_start)
    worksheet["H12"] = _excel_date(period_end)

    for row_index, viaje in enumerate(viajes, start=RICCO_DATA_START_ROW):
        worksheet[f"B{row_index}"] = _excel_date(viaje.fecha)
        worksheet[f"C{row_index}"] = viaje.carta_porte or None
        worksheet[f"D{row_index}"] = viaje.carga or None
        worksheet[f"E{row_index}"] = _ricco_subclient_name(viaje.cliente) or None
        worksheet[f"F{row_index}"] = _ricco_tipo_carga(viaje.tipo_carga) or None
        worksheet[f"G{row_index}"] = viaje.lugar_carga or None
        worksheet[f"H{row_index}"] = viaje.lugar_descarga or None
        worksheet[f"I{row_index}"] = viaje.chofer or None
        worksheet[f"J{row_index}"] = _camion_patente(viaje.camion) or None
        worksheet[f"K{row_index}"] = viaje.tarifa
        worksheet[f"L{row_index}"] = _excel_date(viaje.fecha_descarga_vacio)
        worksheet[f"M{row_index}"] = viaje.demora
        worksheet[f"N{row_index}"] = viaje.costo_total
        worksheet[f"O{row_index}"] = 0
        worksheet[f"P{row_index}"] = viaje.gas_oil_lts
        worksheet[f"Q{row_index}"] = viaje.costo_total
        worksheet[f"R{row_index}"] = viaje.observaciones or None

    worksheet[f"B{total_row}"] = "Total a facturar"
    worksheet[f"K{total_row}"] = _sum_formula("K", viajes, total_row)
    worksheet[f"M{total_row}"] = _sum_formula("M", viajes, total_row)
    worksheet[f"N{total_row}"] = _sum_formula("N", viajes, total_row)
    worksheet[f"O{total_row}"] = _sum_formula("O", viajes, total_row)
    worksheet[f"P{total_row}"] = _sum_formula("P", viajes, total_row)
    worksheet[f"Q{total_row}"] = _sum_formula("Q", viajes, total_row)

    workbook.save(output_path)


def _report_headers() -> list[str]:
    return [
        "Fecha",
        "Cliente",
        "N° Carta de Porte",
        "Carga",
        "Lugar carga",
        "Lugar descarga",
        "Chofer",
        "Camion",
        "Tarifa",
        "Demora",
        "Vacio",
        "Gas oil (lts)",
        "Peajes",
        "Total",
        "Estado",
    ]


def _report_rows(viajes: list[ViajeResumen]) -> list[list[str]]:
    return [
        [
            viaje.fecha,
            viaje.cliente,
            viaje.carta_porte,
            viaje.carga,
            viaje.lugar_carga,
            viaje.lugar_descarga,
            viaje.chofer,
            viaje.camion,
            _format_money(viaje.tarifa),
            _format_money(viaje.demora),
            _format_money(viaje.vacio),
            _format_decimal(viaje.gas_oil_lts),
            _format_money(viaje.peajes),
            _format_money(viaje.costo_total),
            viaje.estado,
        ]
        for viaje in viajes
    ]


def _copy_row_style(worksheet, source_row: int, target_row: int) -> None:
    if target_row == source_row:
        return
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    for column_index in range(2, 19):
        source = worksheet.cell(source_row, column_index)
        target = worksheet.cell(target_row, column_index)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def _clear_row_values(worksheet, row_index: int) -> None:
    for column_index in range(2, 19):
        worksheet.cell(row_index, column_index).value = None


def _safe_filename_component(value: str) -> str:
    cleaned = " ".join((value or "").split())
    for invalid_char in '<>:"/\\|?*':
        cleaned = cleaned.replace(invalid_char, " ")
    cleaned = " ".join(cleaned.split())
    return cleaned.strip() or "sin nombre"


def _internal_ricco_template_path() -> Path | None:
    bundled_candidate = (
        Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parents[2]))
        / "gestion_camiones"
        / "assets"
        / RICCO_TEMPLATE_NAME
    )
    if bundled_candidate.exists():
        return bundled_candidate

    source_candidate = Path(__file__).resolve().parents[1] / "assets" / RICCO_TEMPLATE_NAME
    if source_candidate.exists():
        return source_candidate
    return None


def _format_long_date(value: str) -> str:
    date_value = _excel_date(value)
    if isinstance(date_value, datetime):
        date_value = date_value.date()
    if isinstance(date_value, date):
        month_label = MONTH_NAMES_LOWER[date_value.month - 1]
        return f"{date_value.day:02d} de {month_label} de {date_value.year}"
    return value


def _excel_date(value: str) -> date | str | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return value


def _ricco_subclient_name(cliente_label: str) -> str:
    if cliente_label.endswith(")") and " (" in cliente_label:
        return cliente_label.split(" (", 1)[0].strip()
    return cliente_label.strip()


def _ricco_tipo_carga(tipo_carga: str) -> str:
    normalized = tipo_carga.strip()
    if not normalized or normalized.upper() == "GENERAL":
        return ""
    return normalized


def _camion_patente(camion: str) -> str:
    if " - " in camion:
        return camion.rsplit(" - ", 1)[-1].strip()
    return camion.strip()


def _sum_formula(column: str, viajes: list[ViajeResumen], total_row: int) -> int | str:
    if not viajes:
        return 0
    return f"=SUM({column}{RICCO_DATA_START_ROW}:{column}{total_row - 1})"


def _format_money(value: float) -> str:
    return f"$ {value:,.0f}".replace(",", ".")


def _format_decimal(value: float) -> str:
    return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
